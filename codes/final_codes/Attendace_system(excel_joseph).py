import os
import time
from datetime import datetime
from PIL import Image
from pyzbar.pyzbar import decode
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

# ===========================
# Configuration
# ===========================
EXCEL_FILE = "attendance_log.xlsx"
IMAGES_FOLDER = r"G:\My Drive\qr_images"  # update to your Drive path

# ===========================
# Excel Setup Functions
# ===========================
def initialize_excel():
    """Create Excel file if it doesn't exist with proper headers"""
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"

        # Create headers
        headers = ["Student ID", "Name", "Date", "Time", "Status"]
        sheet.append(headers)

        # Format headers (bold)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        workbook.save(EXCEL_FILE)
        print(f"✓ Created new Excel file: {EXCEL_FILE}")
    else:
        print(f"✓ Excel file already exists: {EXCEL_FILE}")


# ===========================
# QR Code Decoding Functions
# ===========================
def decode_qr_from_image(image_path):
    """
    Decode QR code from image file
    Returns decoded string or None
    """
    try:
        img = Image.open(image_path)
        decoded_objects = decode(img)

        if decoded_objects:
            qr_data = decoded_objects[0].data.decode('utf-8')
            print(f"✓ QR Code decoded: {qr_data}")
            return qr_data
        else:
            print(f"✗ No QR code found in image: {image_path}")
            return None

    except Exception as e:
        print(f"✗ Error decoding QR code: {e}")
        return None


# ===========================
# Data Parsing Functions
# ===========================
def parse_qr_data(qr_data):
    """
    Expected: "StudentID:Name" or "StudentID,Name" or just "StudentID"
    Returns dict or None
    """
    try:
        if ':' in qr_data:
            parts = qr_data.split(':', 1)
        elif ',' in qr_data:
            parts = qr_data.split(',', 1)
        else:
            parts = [qr_data.strip(), "Unknown"]

        student_id = parts[0].strip()
        name = parts[1].strip() if len(parts) > 1 else "Unknown"

        return {'student_id': student_id, 'name': name}
    except Exception as e:
        print(f"✗ Error parsing QR data: {e}")
        return None


# ===========================
# Logging Functions
# ===========================
def log_attendance(student_id, name, status="Present"):
    """
    Append attendance to Excel file, avoid duplicate same-day entries
    Returns True if logged, False otherwise
    """
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active

        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")

        # Check duplicate (same student, same date)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == student_id and row[2] == date_str:
                print(f"⚠ Duplicate: {student_id} already logged today at {row[3]}")
                workbook.close()
                return False

        new_row = [student_id, name, date_str, time_str, status]
        sheet.append(new_row)
        workbook.save(EXCEL_FILE)
        print(f"✓ Logged: {student_id} - {name} at {time_str}")
        return True

    except Exception as e:
        print(f"✗ Error logging attendance: {e}")
        return False


# ===========================
# Main Processing Function
# ===========================
def process_qr_image(image_path):
    """
    Full workflow: decode QR code, parse, and log to excel
    """
    print(f"\n▶ Processing: {image_path}")

    qr_data = decode_qr_from_image(image_path)
    if not qr_data:
        return False

    student_info = parse_qr_data(qr_data)
    if not student_info:
        return False

    return log_attendance(student_info['student_id'], student_info['name'])


# ===========================
# Monitoring Function
# ===========================
def monitor_folder(folder_path):
    """
    Monitor folder for new images and process them
    """
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        print(f"✓ Created folder: {folder_path}")

    processed_files = set()

    print(f"\n👁 Monitoring folder: {folder_path}")
    print("Press Ctrl+C to stop\n")

    try:
        while True:
            files = [f for f in os.listdir(folder_path)
                     if f.lower().endswith(('.png', '.jpg', '.jpeg'))]

            for filename in files:
                if filename not in processed_files:
                    image_path = os.path.join(folder_path, filename)
                    processed = process_qr_image(image_path)
                    if processed:
                        # Optionally delete or move processed files to avoid reprocessing
                        # os.remove(image_path)
                        processed_files.add(filename)
                    else:
                        # If decode failed, still add to processed_files to prevent infinite loop
                        processed_files.add(filename)

            time.sleep(2)

    except KeyboardInterrupt:
        print("\n\n✓ Monitoring stopped")


# ===========================
# Main Entry Point
# ===========================
def main():
    print("=" * 50)
    print("QR Code Attendance Logging System - Google Drive Version")
    print("=" * 50)

    initialize_excel()

    print("\nSelect mode:")
    print("1. Process single image")
    print("2. Monitor folder for new images")

    choice = input("\nEnter choice (1/2): ").strip()

    if choice == "1":
        image_path = input("Enter image path: ").strip()
        if os.path.exists(image_path):
            process_qr_image(image_path)
        else:
            print("✗ File not found:", image_path)

    elif choice == "2":
        monitor_folder(IMAGES_FOLDER)

    else:
        print("✗ Invalid choice")


if __name__ == "__main__":
    main()