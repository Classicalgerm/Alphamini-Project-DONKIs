pip install pyzbar pillow openpyxl

import os
from datetime import datetime
from PIL import Image
from pyzbar.pyzbar import decode
import openpyxl
from openpyxl import Workbook

# ===========================
# Configuration
# ===========================
EXCEL_FILE = "attendance_log.xlsx"
IMAGES_FOLDER = "received_photos"  # Folder where robot saves photos


# ===========================
# Excel Setup Functions
# ===========================
def initialize_excel():
    """Create Excel file if it doesn't exist with proper headers"""
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"

        # Create headers
        headers = ["Student ID", "Name", "Date", "Time", "Status"]
        sheet.append(headers)

        # Format headers (bold)
        for cell in sheet[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        workbook.save(EXCEL_FILE)
        print(f"✓ Created new Excel file: {EXCEL_FILE}")
    else:
        print(f"✓ Excel file already exists: {EXCEL_FILE}")


# ===========================
# QR Code Decoding Functions
# ===========================
def decode_qr_from_image(image_path):
    """
    Decode QR code from image file

    Args:
        image_path: Path to the image file

    Returns:
        Decoded data as string, or None if no QR code found
    """
    try:
        # Open image using PIL
        img = Image.open(image_path)

        # Decode QR codes in the image
        decoded_objects = decode(img)

        if decoded_objects:
            # Get the first QR code's data
            qr_data = decoded_objects[0].data.decode('utf-8')
            print(f"✓ QR Code decoded: {qr_data}")
            return qr_data
        else:
            print("✗ No QR code found in image")
            return None

    except Exception as e:
        print(f"✗ Error decoding QR code: {e}")
        return None


# ===========================
# Data Parsing Functions
# ===========================
def parse_qr_data(qr_data):
    """
    Parse QR code data into student information

    Expected format: "StudentID:Name" or "StudentID,Name"
    Examples: "S12345:John Doe" or "S12345,John Doe"

    Args:
        qr_data: Raw QR code string

    Returns:
        Dictionary with student_id and name, or None if parsing fails
    """
    try:
        # Try colon separator first
        if ':' in qr_data:
            parts = qr_data.split(':', 1)
        # Try comma separator
        elif ',' in qr_data:
            parts = qr_data.split(',', 1)
        # If just ID is provided
        else:
            parts = [qr_data.strip(), "Unknown"]

        student_id = parts[0].strip()
        name = parts[1].strip() if len(parts) > 1 else "Unknown"

        return {
            'student_id': student_id,
            'name': name
        }
    except Exception as e:
        print(f"✗ Error parsing QR data: {e}")
        return None


# ===========================
# Logging Functions
# ===========================
def log_attendance(student_id, name, status="Present"):
    """
    Log attendance record to Excel file

    Args:
        student_id: Student ID from QR code
        name: Student name from QR code
        status: Attendance status (default: "Present")
    """
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active

        # Get current date and time
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")

        # Check for duplicate entry (same student, same day)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == student_id and row[2] == date_str:
                print(f"⚠ Duplicate: {student_id} already logged today at {row[3]}")
                workbook.close()
                return False

        # Append new record
        new_row = [student_id, name, date_str, time_str, status]
        sheet.append(new_row)

        # Save workbook
        workbook.save(EXCEL_FILE)
        print(f"✓ Logged: {student_id} - {name} at {time_str}")
        return True

    except Exception as e:
        print(f"✗ Error logging attendance: {e}")
        return False


# ===========================
# Main Processing Function
# ===========================
def process_qr_image(image_path):
    """
    Complete workflow: decode QR code and log attendance

    Args:
        image_path: Path to the QR code image

    Returns:
        True if successful, False otherwise
    """
    print(f"\n▶ Processing: {image_path}")

    # Step 1: Decode QR code
    qr_data = decode_qr_from_image(image_path)
    if not qr_data:
        return False

    # Step 2: Parse data
    student_info = parse_qr_data(qr_data)
    if not student_info:
        return False

    # Step 3: Log to Excel
    success = log_attendance(
        student_info['student_id'],
        student_info['name']
    )

    return success


# ===========================
# Monitoring Function
# ===========================
def monitor_folder(folder_path):
    """
    Monitor folder for new images and process them

    Args:
        folder_path: Path to folder where robot saves images
    """
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        print(f"✓ Created folder: {folder_path}")

    processed_files = set()

    print(f"\n👁 Monitoring folder: {folder_path}")
    print("Press Ctrl+C to stop\n")

    try:
        while True:
            # List all image files
            files = [f for f in os.listdir(folder_path)
                     if f.lower().endswith(('.png', '.jpg', '.jpeg'))]

            # Process new files
            for filename in files:
                if filename not in processed_files:
                    image_path = os.path.join(folder_path, filename)
                    process_qr_image(image_path)
                    processed_files.add(filename)

            # Wait before checking again
            import time
            time.sleep(2)

    except KeyboardInterrupt:
        print("\n\n✓ Monitoring stopped")


# ===========================
# Main Entry Point
# ===========================
def main():
    """Main function to run the attendance system"""
    print("=" * 50)
    print("QR Code Attendance Logging System - PC Side")
    print("=" * 50)

    # Initialize Excel file
    initialize_excel()

    # Choose mode
    print("\nSelect mode:")
    print("1. Process single image")
    print("2. Monitor folder for new images")

    choice = input("\nEnter choice (1/2): ").strip()

    if choice == "1":
        # Single image mode
        image_path = input("Enter image path: ").strip()
        process_qr_image(image_path)

    elif choice == "2":
        # Monitoring mode
        monitor_folder(IMAGES_FOLDER)

    else:
        print("✗ Invalid choice")


if __name__ == "__main__":
    main()
